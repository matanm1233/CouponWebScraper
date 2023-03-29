from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import copy
from selenium.common.exceptions import NoSuchElementException
import json
import time
import datetime

# set up web driver
driver = webdriver.Chrome(service= Service(ChromeDriverManager().install()))

start = time.time()

def get_coupon_dict():
    coupons = {}
    # get page html
    url = "https://www.southernsavers.com/kroger-weekly-ad-deals/"

    print("Connecting to site...")
    driver.get(url)
    driver.implicitly_wait(10)

    print("Selecting data...")
    # select Ul that contains coupon data
    rawdata = driver.find_element(By.ID, "store-widget")

    rawdata = driver.find_element(By.CSS_SELECTOR, "ul.list")

    # get the list of coupon elements and their respective heading names from the page
    headings = rawdata.find_elements(By.TAG_NAME, "h3")
    categories = rawdata.find_elements(By.CSS_SELECTOR, "ul.category")

    # generate a dictionary which associates the heading names with the ul elements that contain the
    # associated coupons
    i = 0
    # tell the driver to stop waiting so it stops taking literal hours
    driver.implicitly_wait(0)
    for heading in headings:
        # set a key in the coupons dict to an empty list
        coupons[heading.text] = []
        # get list of li tags in the category
        coupons_li = categories[i].find_elements(By.CSS_SELECTOR, "li.item")
        coupons_info = []
        for li in coupons_li:
            # get all the text we want from the coupon, append the dictionary with a list containing the data

            # we must surround the searches with try/except blocks since selenium will throw an error if
            # we find no element
            item, addnlinfo, link, item_math = "", "", "", ""
            try:
                item = li.find_element(By.TAG_NAME, "h4").text
            except NoSuchElementException:
                pass
            try:
                addnlinfo = li.find_element(By.CSS_SELECTOR, "span.q").text
            except NoSuchElementException:
                pass
            try:
                link = li.find_element(By.TAG_NAME, "a").get_attribute("href")
            except NoSuchElementException:
                pass
            try:
                item_math = li.find_element(By.CSS_SELECTOR, "span.item-math").text
            except NoSuchElementException:
                pass
            dictbuffer = {
                "item": item,
                "additional_info": addnlinfo,
                "item_math": item_math
            }
            if link == "":
                dictbuffer["link"] = "No Link"
            else:
                dictbuffer["link"] = f'=HYPERLINK("{link}", "Link")'

            coupons_info.append(dictbuffer)

        # add the dict to the list for that category
        coupons[heading.text] = coupons_info
        i += 1

    # the format of the dict returned by this function is coupons['section name'][int number of item][key (either
    # 'item', 'additional_info', 'link', or 'item_math')]
    coupon_save = open('coupons.json', 'w')
    coupon_save.write(json.dumps(coupons, indent=4))

    return coupons

def make_spreadsheet(coupons):
    print("Creating spreadsheet...")
    workbook = Workbook()
    date = datetime.date.today()
    workbook.save(f"coupons-{date}.xlsx")
    i = 0
    for heading in coupons.keys():
        active_list_of_dicts = coupons[heading]
        # create sheet for each heading
        sheet = workbook.create_sheet(heading, i)

        # add headings
        sheet.append(list(active_list_of_dicts[0].keys()))

        # add rows
        for item_dict in active_list_of_dicts:
            sheet.append(list(item_dict.values()))
        i += 1


    print("Formatting spreadsheet...")
    # format sheet
    for sheet in workbook.sheetnames:
        linkfont = Font(bold=True, color="ff91d2ff")
        bold = Font(bold=True)

        nofill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        itemfill = PatternFill(start_color="9bbb59", end_color="9bbb59", fill_type="solid")
        linkfill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")
        addlinfofill = PatternFill(start_color="cfe2f3", end_color="cfe2f3", fill_type="solid")
        mathfill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid")
        border =Border(left=Side(style='hair'),
                     right=Side(style='hair'),
                     top=Side(style='hair'),
                     bottom=Side(style='hair'))

        center = Alignment(vertical='center', wrapText=True, horizontal='center')

        formats = [
            {
                "column": "A",
                "fill": itemfill,
                "default": "none",
                "alignment": center
            },
            {
                "column": "B",
                "fill": addlinfofill,
                "default": "",
                "alignment": center
            },
            {
                "column": "C",
                "fill": mathfill,
                "default": "",
                "alignment": center
            },
            {
                "column": "D",
                "fill": linkfill,
                "default": "No Link",
                "alignment": center
            }
        ]

        sheet = workbook.get_sheet_by_name(sheet)
        # format width of columns
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 75
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 40

        for col_dict in formats:
            for cell in sheet[col_dict["column"]]:
                cell.fill = col_dict["fill"]
                alignment = copy.copy(cell.alignment)
                cell.alignment = col_dict["alignment"]
                cell.border = border

        for cell in sheet['D']:
            if cell.value != "No Link":
                cell.font = linkfont

        # format other styles
        for cell in sheet[1]:
            cell.font = bold

    workbook.save(f"coupons-{date}.xlsx")

if __name__ == '__main__':
    make_spreadsheet(get_coupon_dict())
    end = time.time()
    print(f" Done in {round(end - start, 2)} seconds")

